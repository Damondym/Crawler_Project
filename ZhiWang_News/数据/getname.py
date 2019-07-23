import openpyxl
wb = openpyxl.load_workbook("F:\python项目\知网爬虫\数据\数据.xlsx")
sheet = wb.get_sheet_by_name('code1')
for row in range(1,557):
    id_SZ=sheet.cell(row=row,column=1).value[0:6]
    business=sheet.cell(row=row,column=2).value
    earliest_time=sheet.cell(row=row,column=5).value
    print(id_SZ+business+str(earliest_time))
    
