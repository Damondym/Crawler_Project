import requests
from bs4 import BeautifulSoup
import re
import openpyxl
import os
import urllib.request
import time
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
import selenium
#下载pdf
def downloadPDF(url):
    try:
        browser.get(url)
        #获得新闻标题
        news_title=browser.find_element_by_class_name("title").text
        #获得作者名字
        news_writer=browser.find_element_by_class_name("author").find_element_by_xpath("span").text
        #获得正文快照
        news_main=browser.find_element_by_id("ChDivSummary").text
        #获得下载连接
        pdf_url=browser.find_element_by_id("pdfDown").get_attribute("href")
        if ifAlert(pdf_url):
            return
        else:
            pdf_point=browser.find_element_by_id("pdfDown")
            browser.execute_script("arguments[0].scrollIntoView(true);", pdf_point)
            WebDriverWait(browser,10).until(
            EC.presence_of_element_located((By.ID, "pdfDown")))
            browser.find_element_by_id("pdfDown").click()
            global count
            count+=1
            #导入excel
            sheet.cell(row=count+1,column=1,value=news_date[0])
            sheet.cell(row=count+1,column=2,value=news_title)
            sheet.cell(row=count+1,column=3,value=news_writer)
            sheet.cell(row=count+1,column=4,value=news_main)
            sheet.cell(row=count+1,column=5,value=news_publish[0])
            sheet.cell(row=count+1,column=6,value=url)
            pdf_name=titleToPDF(news_title,news_writer)
            global news_title_url
            news_title_url.append([pdf_name,url])
            handles = browser.window_handles
            browser.switch_to_window(handles[0])
            if len(handles)>=5:
                stopAndWait(browser)
            else:
                time.sleep(0.5)
        
    except selenium.common.exceptions.NoSuchElementException:
        downloadPDFWithoutAuthor(url)  
    except Exception as e:
        print(repr(e))
        downloadPDF(url)
        
        '''
    url='http://kns.cnki.net/kcms/detail/detail.aspx?dbcode=CCND&filename=LZBR201706270040&dbname=CCNDLAST2017'
    browser.get(url)
    #获得新闻标题
    news_title=browser.find_element_by_class_name("title").text
    #获得作者名字
    news_writer=browser.find_element_by_class_name("author").find_element_by_xpath("span").text
    #获得正文快照
    news_main=browser.find_element_by_id("ChDivSummary").text
    #获得下载连接
    pdf_url=browser.find_element_by_id("pdfDown").get_attribute("href")
    if ifAlert(pdf_url):
        return
    else:
        global count
        count+=1
        #导入excel
        sheet.cell(row=count+1,column=1,value=news_date[0])
        sheet.cell(row=count+1,column=2,value=news_title)
        sheet.cell(row=count+1,column=3,value=news_writer)
        sheet.cell(row=count+1,column=4,value=news_main)
        sheet.cell(row=count+1,column=5,value=news_publish[0])
        sheet.cell(row=count+1,column=6,value=pdf_url)

        browser.find_element_by_id("pdfDown").click()
    '''
    
def downloadPDFWithoutAuthor(url):
    try:
        browser.get(url)
        #获得新闻标题
        news_title=browser.find_element_by_class_name("title").text
        #获得正文快照
        news_main=browser.find_element_by_id("ChDivSummary").text
        #获得下载连接
        pdf_url=browser.find_element_by_id("pdfDown").get_attribute("href")
        if ifAlert(pdf_url):
            return
        else:
            time.sleep(0.5)
            browser.find_element_by_id("pdfDown").click()
            global count
            count+=1
            #导入excel
            sheet.cell(row=count+1,column=1,value=news_date[0])
            sheet.cell(row=count+1,column=2,value=news_title)
            sheet.cell(row=count+1,column=4,value=news_main)
            sheet.cell(row=count+1,column=5,value=news_publish[0])
            sheet.cell(row=count+1,column=6,value=url)
            pdf_name=titleToPDFWithoutAuther(news_title)
            global news_title_url
            news_title_url.append([pdf_name,url])
            handles = browser.window_handles
            browser.switch_to_window(handles[0])
            if len(handles)>=5:
                stopAndWait(browser)
            else:
                time.sleep(0.5)
    except Exception as e:
        print(repr(e))
        downloadPDFWithoutAuthor(url)
        
def downloadPDF_mistake(url):
    try:
        browser.get(url)
        #获得新闻标题
        news_title=browser.find_element_by_class_name("title").text
        #获得作者名字
        news_writer=browser.find_element_by_class_name("author").find_element_by_xpath("span").text
        #获得下载连接
        pdf_url=browser.find_element_by_id("pdfDown").get_attribute("href")
        #if ifAlert(pdf_url):
        if False:
            return
        else:
            time.sleep(0.5)
            browser.find_element_by_id("pdfDown").click()
            pdf_name=titleToPDF(news_title,news_writer)
            global news_title_url
            news_title_url.append([pdf_name,url])
            handles = browser.window_handles
            browser.switch_to_window(handles[0])
            if len(handles)>=5:
                stopAndWait(browser)
            else:
                time.sleep(0.5)
    except selenium.common.exceptions.NoSuchElementException:
        downloadPDF_mistakeWithoutAuthor(url) 
    except Exception as e:
        print(repr(e))
        downloadPDF_mistake(url)

def downloadPDF_mistakeWithoutAuthor(url):
    try:
        browser.get(url)
        #获得新闻标题
        news_title=browser.find_element_by_class_name("title").text
        #获得下载连接
        pdf_url=browser.find_element_by_id("pdfDown").get_attribute("href")
        #if ifAlert(pdf_url):
        if False:
            return
        else:
            time.sleep(0.5)
            browser.find_element_by_id("pdfDown").click()
            pdf_name=titleToPDFWithoutAuther(news_title)
            global news_title_url
            news_title_url.append([pdf_name,url])
            handles = browser.window_handles
            browser.switch_to_window(handles[0])
            if len(handles)>=5:
                stopAndWait(browser)
            else:
                time.sleep(0.5)
    except Exception as e:
        print(repr(e))
        downloadPDF_mistakeWithoutAuthor(url)
    
def stopAndWait(browser):
    try:
        time.sleep(1)
        handles = browser.window_handles
        print(len(handles))
        if len(handles)>=4:
            time.sleep(1)
            handles = browser.window_handles
        if len(handles)>=4:
            time.sleep(1)
            handles = browser.window_handles
        if len(handles)>=4:
            time.sleep(1)
            handles = browser.window_handles
        if len(handles)>=4:
            time.sleep(1)
            handles = browser.window_handles
        if len(handles)==5:
            browser.switch_to_window(handles[2])
            browser.close()
            browser.switch_to_window(handles[1])
            browser.close()
        if len(handles)==4:
            browser.switch_to_window(handles[1])
            browser.close()
        handles = browser.window_handles
        browser.switch_to_window(handles[0])
    except:
        browser.switch_to_window(handles[0])
       
def ifAlert(pdfUrl):
    pdf_page=requests.get(pdfUrl).content
    '''
    if "alert" in str(pdf_page):
        print('return alert '+url)
        return True
    '''
    if "内容保密" in str(pdf_page):
        print('return 内容保密 '+url)
        return True
    else:
        return False

def titleToPDF(title,auther):
    point = re.compile('[《》？！、，。”“·：.]')
    title_noPoint = point.sub('_',title).replace('__','_').replace(' ','  ',10)
    if title_noPoint[-1:] != '_':
        PDF_name=title_noPoint+'_'+auther+'.pdf'
        return PDF_name
    else:
        PDF_name=title_noPoint+auther+'.pdf'
        return PDF_name

def titleToPDFWithoutAuther(title):
    point = re.compile('[《》？！、，。”“·：.]')
    title_noPoint = point.sub('_',title).replace('__','_').replace(' ','  ',10)
    if title_noPoint[-1:] != '_':
        PDF_name=title_noPoint+'_.pdf'
        return PDF_name
    else:
        PDF_name=title_noPoint+'.pdf'
        return PDF_name
        
#翻页url
def newPage():
    global page
    page+=1
    newUrl=url+'&p='+str(page*15)
    return newUrl

#读取每页的内容
def readMainPage(url):
    #访问页面并返回源代码
    data=requests.get(url).content
    news=BeautifulSoup(data,"html.parser")
    #正文内容
    print('已完成爬取第'+str(page+1)+'页内容')
    news_list=news.find('div',attrs={'class':'articles'})
    readDetailPage(news_list)

#读取文章详情页面的信息
def readDetailPage(news_list):
    #读取搜索页面的每一个文章链接并加工
    
    for news_div in news_list.find_all('div',attrs={'class':'wz_tab'}):
        news_time_publish=news_div.find('span',attrs={'class':'year-count'})
        publish_compile=re.compile('title="([\w\n\s\S]*?)"')
        time_compile=re.compile('(20\d{2}[-/]\d{2}[-/]\d{2})')
        global news_date,news_publish
        news_date=re.findall(time_compile,str(news_time_publish))
        news_publish=re.findall(publish_compile,str(news_time_publish))
        global news_year
        news_year=int(news_date[0][0:4])
        if(earliest_year-1<=news_year):
            #读取每个文章对应的链接
            news_url=news_div.find('a',attrs={'target':'_blank'}).get('href')
            #提取链接有效部分
            key_compile=re.compile('http://epub.cnki.net/grid2008/brief/detailj.aspx\?([\w\n\s\W]*)')
            news_url_key=re.findall(key_compile,news_url)
            #加工成新的有效的连接
            news_url_true='http://kns.cnki.net/kcms/detail/detail.aspx?dbcode=CCND&'+news_url_key[0]
            '''
            #读取新链接内容
            data_detail=requests.get(news_url_true).content
            news_detail=BeautifulSoup(data_detail,"html.parser")
            
            #获得题目
            news_title=news_detail.find('div',attrs={'class':'wxTitle'}).find('h2').getText()
            #获得作者名字
            news_writer=news_detail.find('div',attrs={'class':'author'}).getText()
            #获得正文快照
            news_main=news_detail.find('span',attrs={'id':'ChDivSummary'}).getText()
            #获得下载连接
            news_download_pdf=news_detail.find('a',attrs={'class','icon icon-dlpdf'}).get('href')
            
            sheet.cell(row=count+1,column=i,value=count)
            sheet.cell(row=count+1,column=i+1,value=news_title)
            sheet.cell(row=count+1,column=i+2,value=news_writer)
            sheet.cell(row=count+1,column=i+3,value=news_main)
            sheet.cell(row=count+1,column=i+4,value=news_download_caj)
            
            '''
            downloadPDF(news_url_true)
        else:
            break;
    
    

#获得总页面数量
def readPageCount(url):
    try:
        #访问页面并返回源代码
        data=requests.get(url).content
        news=BeautifulSoup(data,"html.parser")
        #导航列表
        side=news.find('div',attrs={'class':'main'}).find('div',attrs={'class':'side'})
        #搜索结果数量
        count_compile=re.compile('报纸全文 \((\d+)\)</a>')
        count=re.findall(count_compile,str(side))
        return int(count[0])
    except:
        os.makedirs('F:\python项目\知网爬虫\PDF\\'+business_id)
        return 0


wb = openpyxl.load_workbook("F:\python项目\知网爬虫\数据\数据.xlsx")
code = wb.get_sheet_by_name('code1')
earliest_year=0
news_year=0
news_date=0
news_publish=0
news_title_url=[[]]
repeat=0
#for row in range(len(os.listdir("F:\python项目\知网爬虫\PDF"))+1,558):
for row in range(565,567):
    business_id=code.cell(row=row,column=1).value[0:6]
    business=code.cell(row=row,column=2).value
    earliest_year=code.cell(row=row,column=5).value
    news_year=earliest_year
    #根据搜索内容制成访问url
    url='http://search.cnki.net/search.aspx?q='+business+'&rank=date&cluster=zyk&val=CCNDTOTAL'
    # 创建Chrome浏览器配置对象实例
    chromeOptions = webdriver.ChromeOptions()
    # 设定下载文件的保存目录为指定目录
    # 如果该目录不存在，将会自动创建
    prefs = {'profile.default_content_setting_values':{'notifications':2},
             "download.default_directory": "F:\python项目\知网爬虫\PDF\\"+business_id}
    # 将自定义设置添加到Chrome配置对象实例中
    chromeOptions.add_experimental_option("prefs", prefs)
    # 启动带有自定义设置的Chrome浏览器
    browser= webdriver.Chrome(chrome_options=chromeOptions)
    page=-1
    count=0
    news_title_url=[[]]
    
    wb=openpyxl.load_workbook('新建 Microsoft Excel 工作表.xlsx')
    sheet=wb.get_sheet_by_name('Sheet1')
    sheet.cell(1,1,value='日期')
    sheet.cell(1,2,value='标题')
    sheet.cell(1,3,value='作者')
    sheet.cell(1,4,value='正文快照')
    sheet.cell(1,5,value='报纸名称')
    sheet.cell(1,6,value='新闻页面链接')
    
    print("正在爬取 "+business)
    search_count=readPageCount(url)
    page_count=(search_count+14)//15
    print(page_count)
    
    while(earliest_year-1<=news_year and page_count>=page+2 and page<=67):
        new_url=newPage()
        readMainPage(new_url)
    time.sleep(3)
    
    if not os.path.exists("F:\python项目\知网爬虫\PDF\\"+business_id):
        os.makedirs('F:\python项目\知网爬虫\PDF\\'+business_id)
    print('开始补漏')
    repeat=0
    while(count>len(os.listdir("F:\python项目\知网爬虫\PDF\\"+business_id)) and repeat<=2):
        for i in range(1,len(news_title_url)):
            if os.path.exists("F:/python项目/知网爬虫/PDF/"+business_id+"/"+news_title_url[1][0]) or os.path.exists("F:/python项目/知网爬虫/PDF/"+business_id+"/"+news_title_url[1][0].replace('  ',' ',10)):
                news_title_url.remove(news_title_url[1])
            else:
                url=news_title_url[1][1]
                news_title_url.remove(news_title_url[1])
                downloadPDF_mistake(url)
        time.sleep(3)
        repeat+=1
                       
    print("完成爬取 "+business+"，已下载"+str(count)+"个PDF文件")
    wb.save("F:\python项目\知网爬虫\PDF\\"+business_id+"\\"+business_id+".xlsx")
    browser.quit()
'''
#根据搜索内容制成访问url
url='http://search.cnki.net/search.aspx?q='+keyWord+'&rank=date&cluster=zyk&val=CCNDTOTAL'
#当前页面
page=-1
#当前数据量
count=0
#搜索结果数量
search_count=readPageCount(url)
#根据需要数量爬取数据
need_count=int(input('含数据量为'+str(search_count)+'条，请输入你需要的数据量：'))
#所需页面数量
need_page=need_count//15+1

#初始化excel
wb=openpyxl.load_workbook('知网数据.xlsx')
sheet=wb.get_sheet_by_name('Sheet1')
sheet.cell(1,1,value='序列号')
sheet.cell(1,2,value='标题')
sheet.cell(1,3,value='作者')
sheet.cell(1,4,value='正文快照')
sheet.cell(1,5,value='CAJ下载链接')

# 创建Chrome浏览器配置对象实例
chromeOptions = webdriver.ChromeOptions()
# 设定下载文件的保存目录为C盘的iDownload目录，
# 如果该目录不存在，将会自动创建
prefs = {"download.default_directory": "F:\python项目\知网爬虫\\"+keyWord}
# 将自定义设置添加到Chrome配置对象实例中
chromeOptions.add_experimental_option("prefs", prefs)
# 启动带有自定义设置的Chrome浏览器
browser= webdriver.Chrome(chrome_options=chromeOptions)

while(page<(need_page)-1):
    new_url=newPage()
    readMainPage(new_url)
time.sleep(3)
browser.quit()

wb.save(keyWord+'  的新闻.xlsx')
print('爬取完成')
'''




